import configparser
import datetime
import getpass
import logging
import os
import sys
import time
import warnings
import urllib.parse as urlparse
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

from bs4 import BeautifulSoup
from halo import Halo
from progress.bar import ChargingBar

from py_jama_rest_client.client import JamaClient, APIException

locked_item_data = dict()

def init_jama_client():
    # do we have credentials in the config?
    credentials_dict = {}
    if 'CREDENTIALS' in config:
        credentials_dict = config['CREDENTIALS']
    try:
        instance_url = get_instance_url(credentials_dict)
        oauth = get_oauth(credentials_dict)
        username = get_username(credentials_dict)
        password = get_password(credentials_dict)
        disable_ssl = get_disable_ssl(credentials_dict)
        jama_client = JamaClient(instance_url, credentials=(username, password), oauth=oauth, verify=not disable_ssl)
        jama_client.get_available_endpoints()
        return jama_client
    except APIException:
        # we cant do things without the API so let's kick out of the execution.
        logger.info('Error: invalid Jama credentials, check they are valid in the config.ini file.')
    except:
        logger.info('Failed to authenticate to <' + get_instance_url(credentials_dict) + '>')

    response = input('\nWould you like to manually enter server credentials?\n')
    response = response.lower()
    if response == 'y' or response == 'yes' or response == 'true':
        config['CREDENTIALS'] = {}
        return init_jama_client()
    else:
        sys.exit()


def get_instance_url(credentials_object):
    if 'instance url' in credentials_object:
        instance_url = str(credentials_object['instance url'])
        instance_url = instance_url.lower()
        # ends with a slash? let's remove this
        if instance_url.endswith('/'):
            instance_url = instance_url[:-1]
        # user forget to put the "https://" bit?
        if not instance_url.startswith('https://') or instance_url.startswith('http://'):
            # if forgotten then ASSuME that this is an https server.
            instance_url = 'https://' + instance_url
        # also allow for shorthand cloud instances
        if '.' not in instance_url:
            instance_url = instance_url + '.jamacloud.com'
        return instance_url
    # otherwise the user did not specify this in the config. prompt the user for it now
    else:
        instance_url = input('Enter the Jama Instance URL:\n')
        credentials_object['instance url'] = instance_url
        return get_instance_url(credentials_object)


def get_username(credentials_object):
    if 'username' in credentials_object:
        username = str(credentials_object['username'])
        return username.strip()
    else:
        username = input('Enter the username (basic auth) or client ID (oAuth):\n')
        credentials_object['username'] = username
        return get_username(credentials_object)


def get_password(credentials_object):
    if 'password' in credentials_object:
        password = str(credentials_object['password'])
        return password.strip()
    else:
        password = getpass.getpass(prompt='Enter your password (basic auth) or client secret (oAuth):\n')
        credentials_object['password'] = password
        return get_password(credentials_object)


def get_oauth(credentials_object):
    if 'using oauth' in credentials_object:
        # this is user input here so let's be extra careful
        user_input = credentials_object['using oauth'].lower()
        user_input = user_input.strip()
        return user_input == 'true' or user_input == 'yes' or user_input == 'y'
    else:
        oauth = input('Using oAuth to authenticate?\n')
        credentials_object['using oauth'] = oauth
        return get_oauth(credentials_object)


def get_disable_ssl(credentials_object):
    # this is an optional param so if not specified then return false
    if 'disable ssl' in credentials_object:
        # this is user input here so let's be extra careful
        user_input = credentials_object['disable ssl'].lower()
        user_input = user_input.strip()
        return user_input == 'true' or user_input == 'yes' or user_input == 'y'
    else:
        return False


def get_link_mode():
    # this field is required
    try:
        user_input = config['PARAMETERS']['link mode'].lower()
        user_input = user_input.strip()
        return user_input == 'true' or user_input == 'yes' or user_input == 'y'
    except:
        logger.error("missing 'link mode' parameter... please provide a project id in the config ini")
        sys.exit()


def get_text_mode():
    # this field is required
    try:
        user_input = config['PARAMETERS']['text mode'].lower()
        user_input = user_input.strip()
        return user_input == 'true' or user_input == 'yes' or user_input == 'y'
    except:
        logger.error("missing 'text mode' parameter... please provide a project id in the config ini")
        sys.exit()


def get_display_attribute():
    # this parameter is optional, defaults to documentKey if not specified
    try:
        user_input = config['PARAMETERS']['display attribute']
        return user_input.strip()
    except:
        logger.warning("missing 'display attribute' parameter. using default parmater")
        return 'documentKey'


def get_project_id():
    try:
        return int(config['PARAMETERS']['project id'])
    except:
        logger.error("missing project id... please provide a project id in the config ini")
        sys.exit()


def init_logger():
    # Setup logging
    try:
        os.makedirs('logs')
    except FileExistsError:
        pass

    current_date_time = datetime.datetime.now().strftime('%m-%d-%Y_%H-%M-%S')
    log_file = 'logs/' + str(current_date_time) + '.log'

    logging.basicConfig(filename=log_file, level=logging.INFO,
                        format='%(asctime)s - %(levelname)s - %(message)s',
                        datefmt='%H:%M:%S')

    logger = logging.getLogger()
    logger.addHandler(logging.StreamHandler(sys.stdout))
    return logger


def get_item_field(item_id, field_key):
    try:
        item = client.get_item(item_id)
        return item['fields'][field_key]
    except APIException as e:
        if e.reason is None:
            logger.error('Unable to retrieve name data on item [' + item_id + ']  with exception:' + str(e.reason))
        else:
            logger.error('Unable to retrieve name data on item [' + item_id + ']')

        return None


def get_synced_item(item_id, project_id):
    try:
        synced_items = client.get_items_synceditems(item_id)
    except APIException as e:
        logger.error('Unable to retrieve synced items for item id:[' + str(item_id) + ']. Exception: ' + str(e))
        return None

    if synced_items is None or len(synced_items) == 0:
        logger.error('Unable to find any synced items for original item with ID:[' + item_id + ']')
        return None

    valid_synced_items = []

    for synced_item in synced_items:
        if synced_item['project'] == project_id:
            valid_synced_items.append(synced_item['id'])

    # only should have one valid synced item here
    if len(valid_synced_items) == 1:
        return valid_synced_items[0]
    elif len(valid_synced_items) > 1:
        logger.error('Multiple synced items found item with ID:[' + item_id + ']')
        return None
    else:
        return None


def start_workbook():
    # Create workbook using openpyxl
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write the headers
    sheet['A1'] = "ID"
    sheet['B1'] = "Locked By"
    sheet['C1'] = "URL Link to Item"

    return workbook


def log_locked_items(item_name, locked_by, url):
    # Check for duplicate rows
    if item_name not in locked_item_data:
        locked_item_data[item_name] = [item_name, locked_by, url]
        logger.info("Item {} is locked and was added to the Excel workbook.".format(item_name))


# link fixer script, will identify broken links from old projects, and correct the links
# a link to the past
if __name__ == '__main__':
    warnings.filterwarnings("ignore", category=UserWarning, module='bs4')
    # int some logging ish
    logger = init_logger()
    start_time = time.time()
    workbook = start_workbook()
    sheet = workbook.active

    logger.info('Running link fixer script')

    config = configparser.ConfigParser()
    config.read('config.ini')
    logger.info('Reading in configuration file')

    # make sure we have a mode specified to run on
    if not get_link_mode() and not get_text_mode():
        logger.info('no modes are selected, exiting')
        sys.exit()

    if get_link_mode():
        logger.info('running link mode')
    if get_text_mode():
        logger.info('running text mode')

    client = init_jama_client()
    instance_url = get_instance_url(config['CREDENTIALS'])
    logger.info('Successfully connected to instance: <' + instance_url + '>')

    # grab the required script parameters from the config.ini file.
    project_id = get_project_id()

    # extra data needed for processing
    valid_project_ids = set()

    """
    STEP ZERO - get all the needed meta data to do this work
    """
    spinner_message = 'Retrieving required meta data from instance...'
    spinner = Halo(text=spinner_message, spinner='dots')
    spinner.start()
    project_list = client.get_projects()
    for project in project_list:
        valid_project_id = project.get('id')
        valid_project_ids.add(valid_project_id)
    spinner.stop()

    """
    STEP ONE - get all items from project    
    """
    # let's validate the project id here before continuing
    if project_id not in valid_project_ids:
        logger.error('Invalid project id provided in the config.ini')
        sys.exit()

    spinner_message = 'Retrieving all items from project ID:[' + str(project_id) + ']'
    spinner = Halo(text=spinner_message, spinner='dots')
    spinner.start()
    items = client.get_items(project_id)
    spinner.stop()
    logger.info('Retrieving ' + str(len(items)) + ' items from project ID:[' + str(project_id) + ']')

    """
    STEP TWO - iterate over all the retrieved items and find bad links   
    """
    broken_link_map = {}
    for item in items:
        item_id = item.get('id')
        item_document_key = item.get('documentKey')
        fields = item.get('fields')
        item_url = instance_url + "/perspective.req#/items/" + str(item_id) + "?projectId=" + str(project_id)

        # Getting lock properties, so we don't need to call the API multiple times for Excel logging
        item_lock_properties = item.get('lock')
        item_locked_by = item_lock_properties.get('lockedBy')

        if item_locked_by is None:
            item_locked_by_fullname = "System"
        else:
            locked_by_user = client.get_user(item_locked_by)
            item_locked_by_firstname = locked_by_user.get("firstName")
            item_locked_by_lastname = locked_by_user.get("lastName")
            item_locked_by_fullname = item_locked_by_firstname + " " + item_locked_by_lastname

        for key in fields:
            original_value = fields[key]
            value = fields[key]
            soup = BeautifulSoup(str(value), 'html.parser')
            hyperlinks = soup.find_all('a')
            bad_link_found = False
            bad_link_count = 0

            if len(hyperlinks) > 0:
                logger.info('\nProcessing ' + str(len(hyperlinks)) + ' hyperlinks on item ID:[' + str(
                    item_id) + '] on field name:[' + str(key) + ']')

            counter = 0

            # remove any duplicate entries in this list.
            hyperlinks = list(dict.fromkeys(hyperlinks))

            # iterate over all the hyperlinks
            for hyperlink in hyperlinks:
                counter += 1
                href = hyperlink.get('href')
                parsed_link = urlparse.urlparse(href)
                hyperlink_string = str(hyperlink)

                # we only want to process jama links. let's skip over all the other links
                if parsed_link.hostname is None or parsed_link.hostname not in instance_url:
                    continue

                url_parameters = urlparse.parse_qs(parsed_link.query)
                linked_project_id = None
                linked_item_id = None
                # do we have paramaters from the url? new url formatting from jama that we need to cover
                if not url_parameters:
                    try:
                        # e.g fragment... /items/10140?projectId=77'
                        link_fragment = parsed_link.fragment
                        linked_project_id = link_fragment.split('projectId=')[1]
                        linked_item_id = link_fragment.split('?')[0].split('items/')[1]
                    except Exception as e:
                        logger.error('failed to get url parameters, error:', e)
                        logger.error('unable to identify project and item ids from link <' +
                                     href + '> skipping current link...')
                        continue
                # otherwise assume that this url will have proper params we can access
                else:
                    try:
                        linked_project_id = url_parameters['projectId'][0]
                        linked_item_id = url_parameters['docId'][0]
                    except Exception as e:
                        logger.error('failed to get url parameters, error:', e)
                        logger.error('unable to identify project and item ids from link <' +
                                     href + '> skipping current link...')
                        continue

                logger.info('--- link ' + str(counter) + ' --- Processing link with item ID:[' + str(
                    linked_item_id) + '] and project ID:[' + str(linked_project_id) + ']...')

                try:
                    original_item = client.get_item(item_id)
                except APIException as e:
                    logger.error('Unable to get original data on item ID:[' + str(item_id) + ']. Exception: ' + str(e))

                # let's see if there is a synced item for this link
                corrected_item_id = get_synced_item(linked_item_id, project_id)

                if int(linked_project_id) == int(project_id) and original_item is not None:
                    # we have a valid link here, but do we have a mismatched name?

                    try:
                        target_item = client.get_item(linked_item_id)
                        corrected_item_id = target_item['id']
                    except APIException as e:
                        logger.error(
                            'Unable to get target item data on item ID:[' + str(
                                linked_item_id) + ']. Exception: ' + str(e))

                    # are we running text mode? if so were updating the name
                    if get_text_mode():
                        sourceName = hyperlink_string[hyperlink_string.index('>') + 1:hyperlink_string.index('</a>')]
                        targetName = get_item_field(target_item['id'], get_display_attribute())
                        targetName = targetName.replace('&', '&amp;') #Encode ampersand as $amp; to match sourceName

                        if sourceName == targetName:
                            logger.info("valid link detected. skipping.")
                            continue

                    # otherwise we already have a valid link, quit
                    else:
                        logger.info("valid link detected. skipping.")
                        continue
                elif original_item is None or original_item is {}:
                    logger.error('Unable to find original item ID:[' + item_id + ']')
                    continue
                elif corrected_item_id is None:
                    logger.error('Unable to find synced item, skipping link')
                    continue

                # we must have a single item id before continuing here.
                # also does this project id param not match the current project?
                # if so then this is a bad link
                # there could potentially be more than one bad link per field value. so
                # let's keep track of that.
                if get_link_mode():
                    logger.info(
                        'Identified incorrect link, will update... item ID:[' + str(corrected_item_id) + ']')

                bad_link_found = True
                bad_link_count += 1

                # is text mode enabled? if so then update the link name here
                corrected_item_name = None
                if get_text_mode():
                    corrected_item_name = get_item_field(corrected_item_id, get_display_attribute())
                    # are we running only text mode here?
                    if not get_link_mode():
                        # dont do any redundant updates
                        sourceName = hyperlink_string[hyperlink_string.index('>') + 1:hyperlink_string.index('</a>')]
                        # skip this link if it's already matching here (no work to be done)
                        if sourceName == corrected_item_name:
                            bad_link_found = False
                            continue

                # otherwise text mode is disabled. so use the existing name here.
                else:
                    try:
                        corrected_item_name = hyperlink_string[
                                              hyperlink_string.index('>') + 1:hyperlink_string.index('</a>')]
                    except:
                        logger.error('failed to resolve link name, this link will not update')
                        continue

                    # let's do the work to change the links name to match the new correct item name

                corrected_hyperlink_string = hyperlink_string[
                                             0:hyperlink_string.index('>') + 1] + corrected_item_name + '</a>'

                #  is link mode enabled?
                if get_link_mode():
                    corrected_hyperlink_string = corrected_hyperlink_string.replace(
                        'projectId=' + str(linked_project_id),
                        'projectId=' + str(project_id))
                    corrected_hyperlink_string = corrected_hyperlink_string.replace('docId=' + str(linked_item_id),
                                                                                    'docId=' + str(
                                                                                        corrected_item_id))

                # if we have made it this far then let's go ahead and replace the hyperlink
                if hyperlink_string in value:
                    value = value.replace(hyperlink_string, corrected_hyperlink_string)
                # otherwise we have a character encoding problem here.
                else:
                    start_link = hyperlink_string[0:hyperlink_string.index('>') + 1]
                    end_link = '</a>'

                    start_index = value.index(start_link) + len(start_link)
                    end_index = 0

                    # iterate over the string until we encounter a "<" to get the end_index
                    for i in range(start_index, len(value)):
                        if value[i] == '<':
                            end_index = i
                            break

                    encoded_name = value[start_index:end_index]
                    hyperlink_string = start_link + encoded_name + end_link
                    value = value.replace(hyperlink_string, corrected_hyperlink_string)

                # we have a bad link for this item?
                if bad_link_found:
                    # Before we replace the hyperlink, let's check if it's locked and log it to Excel if so
                    if item_lock_properties['locked']:
                        logger.info("Item was locked and has a broken link.  Logging to Excel File...\n")
                        log_locked_items(str(item_document_key), str(item_locked_by_fullname), str(item_url))

                    # let's build out an object of all the data we care about for patching and logging
                    else:
                        broken_link_data = {
                            'fieldName': key,
                            'newValue': value,
                            'oldValue': original_value,
                            'counter': str(bad_link_count),
                            'itemId': str(item_id),
                            'itemUrl': str(item_url),
                            'itemLockedBy': str(item_locked_by_fullname),
                            'documentKey': str(item_document_key)
                        }
                        broken_list = broken_link_map.get(item_id)
                        if broken_list is None:
                            broken_list = [broken_link_data]
                        else:
                            broken_list.append(broken_link_data)
                        broken_link_map[item_id] = broken_list

    """
    STEP THREE - fix and log all broken hyperlinks
    """
    # use a progress bar here. this can be a very long-running process
    if len(broken_link_map) > 0:
        with ChargingBar('Updating links ', max=len(broken_link_map),
                         suffix='%(percent).1f%% - %(eta)ds') as bar:
            # iterate over the map, and do work
            for item_id, broken_links in broken_link_map.items():

                patch_list = []
                changed_list = []

                logger.info('Updating link(s) on item ID: [' + str(item_id) + ']')

                for b in broken_links:
                    # log out the old and new rich text values.
                    logger_old_value = b.get('oldValue').replace('\n', '\n\t')
                    logger_new_value = b.get('newValue').replace('\n', '\n\t')
                    logger.info(
                        'Field with name [' + b.get('fieldName') + '] contains ' + b.get('counter') +
                        ' link(s) to be updated')

                    payload = {
                        'op': 'replace',
                        'path': '/fields/' + b.get('fieldName'),
                        'value': b.get('newValue')
                    }
                    patch_list.append(payload)

                # let's try and patch this data
                try:
                    client.patch_item(item_id, patch_list)
                    name = b.get('itemId') if b.get('itemId') is not None else "Unknown Item ID"
                    logger.info('Successfully patched item [' + str(name) + ']')
                except APIException as error:
                    if "locked" in str(error):
                        try:
                            log_locked_items(str(b.get('documentKey')), str(b.get('itemLockedBy')),
                                             b.get('itemUrl'))
                            logger.info("Log locked items method successful for Item ID: " + str(b.get('itemId')))
                        except Exception as e:
                            logger.error('Failed to log locked items for [' + str(b.get('itemId')) + ']')
                            logger.error('Error: ' + str(e))
                    else:
                        # Failed to patch
                        logger.error('Failed to patch item [' + str(b.get('itemId')) + ']')
                        logger.error('API exception response: ' + str(error))

                bar.next()
            bar.finish()
            logger.info('updated ' + str(len(broken_link_map)) + ' link(s)')
    else:
        logger.info('There are zero links to be corrected, exiting...')

    for item in locked_item_data:
        sheet.append(locked_item_data[item])

    column_letter = 'C'
    column_index = openpyxl.utils.cell.column_index_from_string(column_letter)
    hyperlink_font = Font(color="0000FF", underline="single")
    for row in sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index):
        cell = row[0]
        cell.hyperlink = cell.value
        cell.font = hyperlink_font
    workbook.save("locked_items.xlsx")

    # were done here
    elapsed_time = '%.2f' % (time.time() - start_time)
    logger.info('total execution time: ' + elapsed_time + ' seconds')
